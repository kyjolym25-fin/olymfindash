"""
올림피아드교육 부서손익_2025년 12월_보고용.xlsx 파서
각 시트에서 2025년 1~12월 데이터를 추출하여 CHM_DATA_25 JS 상수를 생성
"""
import openpyxl, json, sys

FILE = r'D:\AI PJ\올피_손익분석\올림피아드교육 부서손익_2025년 12월_보고용.xlsx'
COLS_2025 = list(range(30, 42))  # 0-indexed: 2025-01 ~ 2025-12

# ── 계정 매핑 ──────────────────────────────────────────────────

REV_MAP = {
    '교습비 매출액': 'rev_tuition', '교습비매출액(직영)': 'rev_tuition',
    '교습비수입': 'rev_tuition',
    '가맹비수입': 'rev_franchise', '가맹비수입(가맹)': 'rev_franchise',
    '도서수입': 'rev_book',
    '온라인컨텐츠수입': 'rev_online',
    '재료비수입': 'rev_material',
    '방과후학교': 'rev_afterschool',
    '임대료수입': 'rev_rent',
}

EXP_MAP = {
    '직원급여': 'exp_salary',
    '상여금': 'exp_bonus',
    '퇴직급여': 'exp_retire', '퇴직금': 'exp_retire',
    '복리후생비': 'exp_welfare',
    '강사비': 'exp_instructor',
    '강사퇴직연금': 'exp_inst_retire',
    '여비교통비': 'exp_travel',
    '기업업무추진비': 'exp_entertain',
    '통신비': 'exp_comm',
    '수도광열비': 'exp_utility',
    '세금과공과금': 'exp_tax',
    '감가상각비': 'exp_deprec',
    '지급임차료': 'exp_lease',
    '수선비': 'exp_repair',
    '보험료': 'exp_insure',
    '운반비': 'exp_ship',
    '교육훈련비': 'exp_training',
    '도서인쇄비': 'exp_print',
    '차량유지비': 'exp_vehicle',
    '사무용품비': 'exp_supplies', '소모품비': 'exp_supplies',
    '지급수수료': 'exp_fee',
    '광고선전비': 'exp_ad',
    '건물관리비': 'exp_facility',
    '무형고정자산상각': 'exp_intangible',
    '차량운행비': 'exp_car_op',
    '프로그램사용료': 'exp_program',
    '청소용역비': 'exp_clean',
    '행사비': 'exp_event',
    '판매장려금': 'exp_incentive',
    '잡비': 'exp_misc',
    '외주용역비': 'exp_outsource',
    '포인트충당금전입': 'exp_provision',
}

EXP_SKIP = {'인건비', '강사료', '기타비용', '소계', '합계', '대손상각비', '경상연구개발비'}

COM_MAP = {
    '직원급여': 'com_salary', '직원급여(학원안분)': 'com_salary',
    '상여금': 'com_bonus', '상여금(학원안분)': 'com_bonus',
    '퇴직급여': 'com_retire', '퇴직금': 'com_retire', '퇴직급여(학원안분)': 'com_retire',
    '복리후생비': 'com_welfare', '복리후생비(학원안분)': 'com_welfare',
}

# ── 시트 파서 ──────────────────────────────────────────────────

def parse_sheet(ws):
    """
    시트에서 2025년 12개월 데이터 추출
    반환: list of 12 dicts (천원 단위)
    """
    rows = list(ws.iter_rows(values_only=True))
    monthly = [{} for _ in range(12)]
    iv_total = [0] * 12

    section = None

    for row in rows:
        row = list(row) + [None] * max(0, 50 - len(row))

        sec_lbl = str(row[1]).strip() if row[1] else ''
        acct_lbl = str(row[3]).strip() if row[3] else ''

        # 섹션 감지
        if 'Ⅰ' in sec_lbl and '매출' in sec_lbl:
            section = 'rev'
        elif 'Ⅱ' in sec_lbl and '판매' in sec_lbl:
            section = 'exp'
        elif 'Ⅳ' in sec_lbl and '공통' in sec_lbl:
            # Ⅳ 총액 행 기록
            for mi, ci in enumerate(COLS_2025):
                v = row[ci]
                try: iv_total[mi] = round(float(v) / 1000) if v else 0
                except: iv_total[mi] = 0
            section = 'common'
            continue
        elif 'Ⅲ' in sec_lbl or 'Ⅴ' in sec_lbl:
            section = None

        if section is None or not acct_lbl:
            continue

        # 계정 키 매핑
        if section == 'rev':
            key = REV_MAP.get(acct_lbl)
        elif section == 'exp':
            if acct_lbl in EXP_SKIP:
                continue
            key = EXP_MAP.get(acct_lbl)
        elif section == 'common':
            key = COM_MAP.get(acct_lbl)
        else:
            key = None

        if not key:
            continue

        for mi, ci in enumerate(COLS_2025):
            v = row[ci]
            try: val = round(float(v) / 1000) if v else 0
            except: val = 0
            monthly[mi][key] = monthly[mi].get(key, 0) + val

    # com_other = Ⅳ 총액 - 파싱된 com_* 항목 합계
    for mi in range(12):
        tracked = sum(monthly[mi].get(k, 0) for k in ['com_salary', 'com_bonus', 'com_retire', 'com_welfare'])
        com_other = iv_total[mi] - tracked
        if com_other != 0:
            monthly[mi]['com_other'] = com_other

    return monthly


def subtract_monthly(base, *others):
    """base에서 others를 뺀 월별 데이터 반환"""
    result = []
    for mi in range(12):
        d = dict(base[mi])
        for other in others:
            for k, v in other[mi].items():
                d[k] = d.get(k, 0) - v
        result.append(d)
    return result


# ── 메인 파싱 ──────────────────────────────────────────────────

print("Loading workbook...", file=sys.stderr)
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)

# 직접 파싱할 노드 → 시트명
LEAF_MAP = {
    'gwj-m': '광진',
    'gwj-e': '광진영어',
    'sdg-m': '성동',
    'sdg-e': '성동영어',
    'ddm-m': '동대문',
    'ddm-e': '동대문영어',
    'jrl-m': '중랑',
    'jrl-e': '중랑영어',
    'jgy':   '중계',
    'spb':   '송파방이',
    'msa':   '미사',
    'uir':   '위례',
    'eng':   '가맹사업',
    'bghu-c':'방과후사업본부',
    'mirae-c':'미래교육사업부(솔루션, 연구소)',
    'mgmt-c':'경영기획실',
}

# 25년 한시적 캠퍼스 시트 (hak-b 계산용, CHM_DATA_25에는 포함 안 함)
# 학원사업본부 소속: 강동고덕, 송파잠실, 하베르만대치
# U2M사업본부 소속: 강북, 강서발산 (아래 EXTRA_U2M_SHEETS에서 처리)
EXTRA_HAK_SHEETS = ['강동고덕', '송파잠실', '하베르만대치']
EXTRA_U2M_SHEETS = ['강북', '강서발산']

data = {}

# 리프 노드 파싱
for node_id, sname in LEAF_MAP.items():
    print(f"  Parsing {sname} → {node_id}", file=sys.stderr)
    ws = wb[sname]
    data[node_id] = parse_sheet(ws)

# 학원사업본부 총합 파싱 (hak-b 도출용)
print("  Parsing 학원사업본부 (for hak-b)", file=sys.stderr)
hak_total = parse_sheet(wb['학원사업본부'])

# 25년 한시적 캠퍼스 파싱
extra_hak = []
for sname in EXTRA_HAK_SHEETS:
    if sname in wb.sheetnames:
        print(f"  Parsing extra campus: {sname}", file=sys.stderr)
        extra_hak.append(parse_sheet(wb[sname]))

# hak-b = 학원사업본부 - 8개 정규 캠퍼스
# (강동고덕, 송파잠실, 하베르만대치는 별도 노드 없이 hak-b에 포함)
hak_campuses = [data[n] for n in ['gwj-m','gwj-e','sdg-m','sdg-e','ddm-m','ddm-e','jrl-m','jrl-e']]
hak_b_monthly = subtract_monthly(hak_total, *hak_campuses)
data['hak-b'] = hak_b_monthly

# U2M사업본부 총합 파싱 (u2m-b 도출용)
print("  Parsing U2M사업본부 (for u2m-b)", file=sys.stderr)
u2m_total = parse_sheet(wb['U2M사업본부'])
# (강북, 강서발산은 별도 노드 없이 u2m-b에 포함)
u2m_campuses = [data[n] for n in ['jgy','spb','msa','uir','eng']]
u2m_b_monthly = subtract_monthly(u2m_total, *u2m_campuses)
data['u2m-b'] = u2m_b_monthly

# strat (전략기획실): 별도 시트 없음 → 0
data['strat'] = [{} for _ in range(12)]

# ── JS 출력 ──────────────────────────────────────────────────

# 0값 제거하여 용량 절감
def clean(d):
    return {k: v for k, v in d.items() if v != 0}

result = {}
for nid, monthly in data.items():
    result[nid] = [clean(m) for m in monthly]

js_str = 'const CHM_DATA_25 = ' + json.dumps(result, ensure_ascii=False, separators=(',', ':')) + ';'
print(js_str)
print(f"\n// Total nodes: {len(result)}", file=sys.stderr)
print("Done.", file=sys.stderr)
